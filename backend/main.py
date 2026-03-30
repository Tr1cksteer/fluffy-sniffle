"""
ЛогоперРадар — Fleet Tracking Backend
FastAPI + SQLite + APScheduler
Run: uvicorn main:app --host 0.0.0.0 --port 8000
"""

import os
import re
import json
import logging
import secrets
import hashlib
from datetime import datetime, timedelta
from typing import Optional, List
from pathlib import Path

from fastapi import FastAPI, HTTPException, Request, Response, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from markupsafe import Markup
from pydantic import BaseModel
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

from database import init_db, get_db_conn
from scraper import fetch_vessel_info
from basin import determine_basin

# ─── Config ──────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent.parent
ADMIN_PASSWORD_HASH = hashlib.sha256(
    os.environ.get("ADMIN_PASSWORD", "logoper2024").encode()
).hexdigest()
SESSION_STORE: dict[str, dict] = {}
SESSION_TTL_HOURS = 12

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ─── App ─────────────────────────────────────────────────────────────────────

app = FastAPI(title="ЛогоперРадар", version="1.0.0")

static_dir = BASE_DIR / "frontend" / "static"
static_dir.mkdir(parents=True, exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

templates = Jinja2Templates(directory=str(BASE_DIR / "frontend" / "templates"))


# ─── Jinja2 filters ──────────────────────────────────────────────────────────

def basin_dot_filter(basin: str) -> Markup:
    if not basin or basin == "—":
        return Markup('<span class="basin-dot unknown">—</span>')
    b = basin.lower()
    if "дв без" in b:
        css = "dv-norf"
    elif "дв каботаж" in b:
        css = "dv-kab"
    elif "дв" in b:
        css = "dv"
    elif "балтика каботаж" in b or ("балт" in b and "каботаж" in b):
        css = "blt-kab"
    elif "балт" in b:
        css = "blt"
    elif "новор" in b:
        css = "nvr"
    elif "транзит" in b:
        css = "transit"
    else:
        css = "unknown"
    safe = Markup.escape(basin)
    return Markup(f'<span class="basin-dot {css}">{safe}</span>')


def line_badge_filter(line: str) -> Markup:
    if not line:
        return Markup('<span class="badge badge-default">• —</span>')
    l = line.lower()
    if "транзит" in l:
        css = "badge-transit"
    elif "fesco" in l:
        css = "badge-fesco"
    elif "регул" in l or "regular" in l:
        css = "badge-reg"
    else:
        css = "badge-default"
    safe = Markup.escape(line)
    return Markup(f'<span class="badge {css}">• {safe}</span>')


templates.env.filters["basinDot"] = basin_dot_filter
templates.env.filters["lineBadge"] = line_badge_filter


# ─── Auth helpers ────────────────────────────────────────────────────────────

def create_session() -> str:
    token = secrets.token_hex(32)
    SESSION_STORE[token] = {"expires": datetime.utcnow() + timedelta(hours=SESSION_TTL_HOURS)}
    return token


def validate_session(token: Optional[str]) -> bool:
    if not token or token not in SESSION_STORE:
        return False
    sess = SESSION_STORE[token]
    if datetime.utcnow() > sess["expires"]:
        del SESSION_STORE[token]
        return False
    sess["expires"] = datetime.utcnow() + timedelta(hours=SESSION_TTL_HOURS)
    return True


def get_session_token(request: Request) -> Optional[str]:
    return request.cookies.get("session")


def require_auth(request: Request):
    if not validate_session(get_session_token(request)):
        raise HTTPException(status_code=401, detail="Unauthorized")


# ─── Schemas ─────────────────────────────────────────────────────────────────

class LoginBody(BaseModel):
    password: str

class AddVesselBody(BaseModel):
    imo: str

class DeleteVesselsBody(BaseModel):
    imos: List[str]


# ─── Scheduler ───────────────────────────────────────────────────────────────

scheduler = AsyncIOScheduler()


@app.on_event("startup")
async def startup():
    init_db()
    scheduler.add_job(
        scheduled_update,
        IntervalTrigger(days=3),
        id="fleet_update",
        replace_existing=True,
        next_run_time=None,
    )
    scheduler.start()
    log.info("App started. Fleet update every 3 days.")


@app.on_event("shutdown")
async def shutdown():
    scheduler.shutdown(wait=False)


async def scheduled_update():
    log.info("Scheduled fleet update started")
    conn = get_db_conn()
    try:
        rows = conn.execute("SELECT imo FROM vessels").fetchall()
        for row in rows:
            await _refresh_vessel_internal(row["imo"], conn)
        conn.execute("UPDATE meta SET value=? WHERE key='last_update'",
                     (datetime.now().strftime("%d.%m.%Y %H:%M"),))
        conn.commit()
    finally:
        conn.close()
    log.info("Scheduled fleet update done")


async def _refresh_vessel_internal(imo: str, conn):
    try:
        info = await fetch_vessel_info(imo)
        if info:
            basin = determine_basin(info.get("ports", []), info.get("route_ports", []))
            conn.execute(
                "UPDATE vessels SET name=?,line=?,current_port=?,destination=?,basin=?,last_seen=?,raw_json=?,updated_at=? WHERE imo=?",
                (
                    info.get("name", ""),
                    info.get("line", ""),
                    info.get("current_port", ""),
                    info.get("destination", ""),
                    basin,
                    info.get("last_seen", ""),
                    json.dumps(info, ensure_ascii=False),
                    datetime.now().strftime("%d.%m.%Y %H:%M"),
                    imo,
                ),
            )
    except Exception as e:
        log.error(f"Refresh error IMO {imo}: {e}")


async def refresh_vessel(imo: str):
    conn = get_db_conn()
    try:
        await _refresh_vessel_internal(imo, conn)
        conn.commit()
    finally:
        conn.close()


# ─── Pages ───────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def public_page(request: Request):
    if validate_session(get_session_token(request)):
        return RedirectResponse("/admin", status_code=302)
    conn = get_db_conn()
    try:
        meta = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM meta").fetchall()}
        stats = _get_basin_stats(conn)
    finally:
        conn.close()
    return templates.TemplateResponse("public.html", {
        "request": request,
        "last_update": meta.get("last_update", "—"),
        "stats": stats,
    })


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    if not validate_session(get_session_token(request)):
        return RedirectResponse("/?login=1", status_code=302)
    conn = get_db_conn()
    try:
        meta = {r["key"]: r["value"] for r in conn.execute("SELECT key, value FROM meta").fetchall()}
        vessels = conn.execute(
            "SELECT imo, name, line, basin, current_port, destination, last_seen, updated_at FROM vessels ORDER BY name"
        ).fetchall()
        stats = _get_basin_stats(conn)
    finally:
        conn.close()
    return templates.TemplateResponse("admin.html", {
        "request": request,
        "vessels": [dict(v) for v in vessels],
        "last_update": meta.get("last_update", "—"),
        "stats": stats,
    })


# ─── API ─────────────────────────────────────────────────────────────────────

@app.post("/api/login")
async def api_login(body: LoginBody, response: Response):
    h = hashlib.sha256(body.password.encode()).hexdigest()
    if h != ADMIN_PASSWORD_HASH:
        raise HTTPException(status_code=403, detail="Неверный пароль")
    token = create_session()
    response.set_cookie("session", token, httponly=True, samesite="lax",
                        max_age=SESSION_TTL_HOURS * 3600)
    return {"ok": True}


@app.post("/api/logout")
async def api_logout(request: Request, response: Response):
    token = get_session_token(request)
    if token in SESSION_STORE:
        del SESSION_STORE[token]
    response.delete_cookie("session")
    return {"ok": True}


@app.get("/api/vessels")
async def api_list_vessels(request: Request):
    require_auth(request)
    conn = get_db_conn()
    try:
        rows = conn.execute(
            "SELECT imo, name, line, basin, current_port, destination, last_seen, updated_at FROM vessels ORDER BY name"
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.post("/api/vessels")
async def api_add_vessel(body: AddVesselBody, request: Request):
    require_auth(request)
    imo = body.imo.strip()
    if not re.match(r"^\d{7}$", imo):
        raise HTTPException(status_code=400, detail="IMO должен содержать ровно 7 цифр")
    conn = get_db_conn()
    try:
        if conn.execute("SELECT imo FROM vessels WHERE imo=?", (imo,)).fetchone():
            raise HTTPException(status_code=409, detail="IMO уже добавлен")
        conn.execute(
            "INSERT INTO vessels (imo,name,line,basin,current_port,destination,last_seen,updated_at) VALUES (?,?,?,?,?,?,?,?)",
            (imo, "", "", "Неизвестно", "", "", "", ""),
        )
        conn.commit()
    finally:
        conn.close()
    import asyncio
    asyncio.create_task(_bg_refresh(imo))
    return {"ok": True, "imo": imo}


async def _bg_refresh(imo: str):
    await refresh_vessel(imo)
    conn = get_db_conn()
    try:
        conn.execute("UPDATE meta SET value=? WHERE key='last_update'",
                     (datetime.now().strftime("%d.%m.%Y %H:%M"),))
        conn.commit()
    finally:
        conn.close()


@app.post("/api/vessels/import")
async def api_import(request: Request, file: UploadFile = File(...)):
    require_auth(request)
    text = (await file.read()).decode("utf-8", errors="ignore")
    imos = list(set(re.findall(r"\b\d{7}\b", text)))
    if not imos:
        raise HTTPException(status_code=400, detail="Не найдено IMO-кодов в файле")
    conn = get_db_conn()
    added, skipped = [], []
    try:
        for imo in imos:
            if conn.execute("SELECT imo FROM vessels WHERE imo=?", (imo,)).fetchone():
                skipped.append(imo)
            else:
                conn.execute(
                    "INSERT INTO vessels (imo,name,line,basin,current_port,destination,last_seen,updated_at) VALUES (?,?,?,?,?,?,?,?)",
                    (imo, "", "", "Неизвестно", "", "", "", ""),
                )
                added.append(imo)
        conn.commit()
    finally:
        conn.close()
    import asyncio
    for imo in added:
        asyncio.create_task(_bg_refresh(imo))
    return {"ok": True, "added": len(added), "skipped": len(skipped)}


@app.delete("/api/vessels/{imo}")
async def api_delete_vessel(imo: str, request: Request):
    require_auth(request)
    conn = get_db_conn()
    try:
        conn.execute("DELETE FROM vessels WHERE imo=?", (imo,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.post("/api/vessels/delete-bulk")
async def api_delete_bulk(body: DeleteVesselsBody, request: Request):
    require_auth(request)
    conn = get_db_conn()
    try:
        for imo in body.imos:
            conn.execute("DELETE FROM vessels WHERE imo=?", (imo,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "deleted": len(body.imos)}


@app.post("/api/vessels/delete-by-file")
async def api_delete_by_file(request: Request, file: UploadFile = File(...)):
    require_auth(request)
    text = (await file.read()).decode("utf-8", errors="ignore")
    imos = list(set(re.findall(r"\b\d{7}\b", text)))
    conn = get_db_conn()
    deleted = 0
    try:
        for imo in imos:
            deleted += conn.execute("DELETE FROM vessels WHERE imo=?", (imo,)).rowcount
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "deleted": deleted}


@app.post("/api/refresh")
async def api_refresh_all(request: Request):
    require_auth(request)
    import asyncio
    asyncio.create_task(_full_refresh())
    return {"ok": True}


async def _full_refresh():
    conn = get_db_conn()
    try:
        imos = [r["imo"] for r in conn.execute("SELECT imo FROM vessels").fetchall()]
    finally:
        conn.close()
    for imo in imos:
        await refresh_vessel(imo)
    conn = get_db_conn()
    try:
        conn.execute("UPDATE meta SET value=? WHERE key='last_update'",
                     (datetime.now().strftime("%d.%m.%Y %H:%M"),))
        conn.commit()
    finally:
        conn.close()


@app.post("/api/refresh/{imo}")
async def api_refresh_single(imo: str, request: Request):
    require_auth(request)
    await refresh_vessel(imo)
    conn = get_db_conn()
    try:
        row = conn.execute(
            "SELECT imo,name,line,basin,current_port,destination,last_seen,updated_at FROM vessels WHERE imo=?", (imo,)
        ).fetchone()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=404)
    return dict(row)


@app.get("/api/stats")
async def api_stats():
    conn = get_db_conn()
    try:
        meta = {r["key"]: r["value"] for r in conn.execute("SELECT key,value FROM meta").fetchall()}
        return {"stats": _get_basin_stats(conn), "last_update": meta.get("last_update", "—")}
    finally:
        conn.close()


@app.get("/api/export/xls")
async def api_export_xls(request: Request):
    require_auth(request)
    conn = get_db_conn()
    try:
        rows = conn.execute(
            "SELECT imo,name,line,basin,current_port,destination,last_seen,updated_at FROM vessels ORDER BY name"
        ).fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Флот"
    hdr_fill = PatternFill("solid", fgColor="0F1C33")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    headers = ["IMO", "Название судна", "Линия", "Бассейн", "Текущий порт", "Назначение", "Обновлено"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(rows, 2):
        ws.cell(ri, 1, row["imo"])
        ws.cell(ri, 2, row["name"] or "—")
        ws.cell(ri, 3, row["line"] or "—")
        ws.cell(ri, 4, row["basin"] or "—")
        ws.cell(ri, 5, row["current_port"] or "—")
        ws.cell(ri, 6, row["destination"] or "—")
        ws.cell(ri, 7, row["updated_at"] or "—")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 40
        )
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fn = f"fleet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'})


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _get_basin_stats(conn) -> list:
    BASINS = ["ДВ", "ДВ каботаж", "ДВ без РФ", "Балтийский", "Балтика каботаж",
              "Новороссийск", "Транзит", "Неизвестно"]
    rows = conn.execute("SELECT basin, COUNT(*) cnt FROM vessels GROUP BY basin").fetchall()
    counts = {r["basin"]: r["cnt"] for r in rows}
    total = sum(counts.values()) or 1
    result, shown = [], set()
    for basin in BASINS:
        c = counts.get(basin, 0)
        result.append({"basin": basin, "count": c, "pct": round(c / total * 100)})
        shown.add(basin)
    for basin, cnt in counts.items():
        if basin not in shown:
            result.append({"basin": basin, "count": cnt, "pct": round(cnt / total * 100)})
    return result
